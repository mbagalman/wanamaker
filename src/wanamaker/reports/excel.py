"""Excel workbook export — analyst-facing structured tables.

The HTML showcase and Markdown report are designed for *reading*; the
Excel workbook is designed for *slicing*. Analysts often want to pivot
channel contributions, plug ROI estimates into their own spreadsheets,
or attach a structured table to a board deck. The workbook gives them
that without re-typing.

Per AGENTS.md Hard Rule 2: **no LLM calls for output generation**, ever.
The workbook is built from the same posterior summary the other
report surfaces consume.

Public surface:

- ``build_excel_workbook`` — turn ``PosteriorSummary`` + ``TrustCard``
  (plus optional run metadata, refresh diff, scenario forecasts) into
  an in-memory ``openpyxl.Workbook`` ready to be saved.
- ``write_excel_workbook`` — convenience wrapper that builds and saves
  to a path in one call.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook

from wanamaker.engine.summary import PosteriorSummary
from wanamaker.forecast.posterior_predictive import ForecastResult
from wanamaker.refresh.diff import RefreshDiff
from wanamaker.trust_card.card import TrustCard, TrustStatus


@dataclass(frozen=True)
class WorkbookMetadata:
    """Run-level metadata included in the Summary sheet header."""

    run_id: str
    period_start: str
    period_end: str
    n_periods: int
    generated_at: datetime
    package_version: str
    runtime_mode: str = "(unknown)"
    engine_label: str = "(unknown)"
    run_fingerprint: str = ""


_HEADER_FILL = "EEF4F8"
_HEADER_FONT_BOLD = True


def _bold(cell: Any) -> None:
    from openpyxl.styles import Font

    cell.font = Font(bold=True)


def _autosize_columns(ws: Any) -> None:
    """Pick a sensible column width so the workbook opens readable.

    openpyxl doesn't auto-size; we estimate the max content width per
    column and clamp to ``[10, 60]`` characters.
    """
    for column_cells in ws.columns:
        if not column_cells:
            continue
        column_letter = column_cells[0].column_letter
        longest = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > longest:
                longest = length
        ws.column_dimensions[column_letter].width = min(max(longest + 2, 10), 60)


def _verdict_label(card: TrustCard) -> str:
    """Worst-dimension verdict, mirroring the showcase pill logic."""
    if not card.dimensions:
        return "no verdict"
    statuses = {d.status for d in card.dimensions}
    if TrustStatus.WEAK in statuses:
        return "weak"
    if TrustStatus.MODERATE in statuses:
        return "moderate"
    return "pass"


def _write_summary_sheet(
    wb: Workbook,
    summary: PosteriorSummary,
    trust_card: TrustCard,
    metadata: WorkbookMetadata,
) -> None:
    ws = wb.active
    ws.title = "Summary"

    rows: list[tuple[str, Any]] = [
        ("Wanamaker MMM Summary", ""),
        ("", ""),
        ("Run ID", metadata.run_id),
        ("Run fingerprint", metadata.run_fingerprint or "—"),
        ("Period", f"{metadata.period_start} – {metadata.period_end}"),
        ("Periods covered", metadata.n_periods),
        (
            "Generated (UTC)",
            metadata.generated_at.strftime("%Y-%m-%d %H:%M"),
        ),
        ("Engine", metadata.engine_label),
        ("Runtime", metadata.runtime_mode),
        ("Package version", metadata.package_version),
        ("", ""),
    ]

    total_media = sum(c.mean_contribution for c in summary.channel_contributions)
    rows.extend([
        ("Total media contribution", float(total_media)),
        ("Trust Card verdict", _verdict_label(trust_card)),
    ])
    if summary.channel_contributions:
        ranked = sorted(
            summary.channel_contributions,
            key=lambda c: c.mean_contribution,
            reverse=True,
        )
        top = ranked[0]
        rows.append((
            "Top channel by contribution",
            f"{top.channel} ({top.mean_contribution / total_media:.0%} of media)"
            if total_media > 0
            else top.channel,
        ))

    for idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        # Bold either the title row (row 1) or any labeled metadata row from
        # row 3 onwards; row 2 is intentionally a blank spacer.
        if label and ((not value and idx == 1) or idx >= 3):
            _bold(ws.cell(row=idx, column=1))

    _autosize_columns(ws)


def _write_channels_sheet(wb: Workbook, summary: PosteriorSummary) -> None:
    ws = wb.create_sheet("Channels")

    headers = [
        "Channel",
        "Mean contribution",
        "HDI low",
        "HDI high",
        "Share of media",
        "ROI (mean)",
        "ROI HDI low",
        "ROI HDI high",
        "Spend min (observed)",
        "Spend max (observed)",
        "Spend invariant",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _bold(cell)

    total_media = sum(c.mean_contribution for c in summary.channel_contributions)
    ranked = sorted(
        summary.channel_contributions,
        key=lambda c: c.mean_contribution,
        reverse=True,
    )
    for r, channel in enumerate(ranked, start=2):
        share = (
            channel.mean_contribution / total_media if total_media > 0 else 0.0
        )
        values = [
            channel.channel,
            float(channel.mean_contribution),
            float(channel.hdi_low),
            float(channel.hdi_high),
            float(share),
            float(channel.roi_mean) if not channel.spend_invariant else None,
            float(channel.roi_hdi_low) if not channel.spend_invariant else None,
            float(channel.roi_hdi_high) if not channel.spend_invariant else None,
            float(channel.observed_spend_min),
            float(channel.observed_spend_max),
            "yes" if channel.spend_invariant else "no",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=value)

    _autosize_columns(ws)


def _write_trust_card_sheet(wb: Workbook, trust_card: TrustCard) -> None:
    ws = wb.create_sheet("Trust Card")

    headers = ["Dimension", "Status", "Explanation"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _bold(cell)

    if not trust_card.dimensions:
        ws.cell(row=2, column=1, value="(no dimensions computed)")
        _autosize_columns(ws)
        return

    for r, dim in enumerate(trust_card.dimensions, start=2):
        ws.cell(row=r, column=1, value=dim.name)
        ws.cell(row=r, column=2, value=dim.status.value)
        ws.cell(row=r, column=3, value=dim.explanation)

    _autosize_columns(ws)


def _write_parameters_sheet(wb: Workbook, summary: PosteriorSummary) -> None:
    if not summary.parameters:
        return
    ws = wb.create_sheet("Parameters")

    headers = ["Parameter", "Mean", "SD", "HDI low", "HDI high", "R-hat", "ESS bulk"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _bold(cell)

    for r, param in enumerate(summary.parameters, start=2):
        ws.cell(row=r, column=1, value=param.name)
        ws.cell(row=r, column=2, value=float(param.mean))
        ws.cell(row=r, column=3, value=float(param.sd))
        ws.cell(row=r, column=4, value=float(param.hdi_low))
        ws.cell(row=r, column=5, value=float(param.hdi_high))
        ws.cell(row=r, column=6, value=float(param.r_hat) if param.r_hat is not None else None)
        ws.cell(
            row=r, column=7,
            value=float(param.ess_bulk) if param.ess_bulk is not None else None,
        )

    _autosize_columns(ws)


def _write_refresh_sheet(wb: Workbook, refresh_diff: RefreshDiff) -> None:
    ws = wb.create_sheet("Refresh diff")

    headers = [
        "Parameter",
        "Previous mean",
        "Current mean",
        "Previous CI low",
        "Previous CI high",
        "Current CI low",
        "Current CI high",
        "Movement class",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _bold(cell)

    for r, m in enumerate(refresh_diff.movements, start=2):
        ws.cell(row=r, column=1, value=m.name)
        ws.cell(row=r, column=2, value=float(m.previous_mean))
        ws.cell(row=r, column=3, value=float(m.current_mean))
        ws.cell(row=r, column=4, value=float(m.previous_ci[0]))
        ws.cell(row=r, column=5, value=float(m.previous_ci[1]))
        ws.cell(row=r, column=6, value=float(m.current_ci[0]))
        ws.cell(row=r, column=7, value=float(m.current_ci[1]))
        ws.cell(
            row=r, column=8,
            value=m.movement_class.value if m.movement_class is not None else "",
        )

    _autosize_columns(ws)


def _write_scenarios_sheet(
    wb: Workbook,
    forecasts: list[ForecastResult],
    plan_names: list[str],
) -> None:
    if not forecasts:
        return
    ws = wb.create_sheet("Scenarios")

    headers = [
        "Plan",
        "Predicted total",
        "Total HDI low",
        "Total HDI high",
        "Δ vs first plan",
        "Extrapolation flags",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _bold(cell)

    means = [float(sum(f.mean)) for f in forecasts]
    baseline_total = means[0] if means else 0.0
    for r, (name, forecast, mean_total) in enumerate(
        zip(plan_names, forecasts, means, strict=True), start=2,
    ):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=mean_total)
        ws.cell(row=r, column=3, value=float(sum(forecast.hdi_low)))
        ws.cell(row=r, column=4, value=float(sum(forecast.hdi_high)))
        ws.cell(
            row=r, column=5,
            value=mean_total - baseline_total if r > 2 else 0.0,
        )
        ws.cell(row=r, column=6, value=len(forecast.extrapolation_flags))

    _autosize_columns(ws)


def build_excel_workbook(
    summary: PosteriorSummary,
    trust_card: TrustCard,
    *,
    metadata: WorkbookMetadata,
    refresh_diff: RefreshDiff | None = None,
    scenarios: list[ForecastResult] | None = None,
    scenario_plan_names: list[str] | None = None,
) -> Workbook:
    """Build the analyst-facing Excel workbook in memory.

    Sheets:
    - **Summary** — run metadata, headline numbers, Trust Card verdict.
    - **Channels** — one row per channel with contribution + ROI + observed
      spend range + spend-invariant flag, ranked by mean contribution.
    - **Trust Card** — one row per dimension (name, status, explanation).
    - **Parameters** — one row per posterior parameter (mean, SD, HDI,
      R-hat, ESS bulk). Skipped when no parameters are available.
    - **Refresh diff** — one row per parameter movement (only when a
      refresh diff is supplied).
    - **Scenarios** — one row per forecast (only when scenarios are
      supplied), with delta vs the first plan.

    Returns the in-memory ``Workbook``; pair with ``Workbook.save(path)``
    or use ``write_excel_workbook`` to combine in one call.
    """
    from openpyxl import Workbook as _Workbook

    wb = _Workbook()
    _write_summary_sheet(wb, summary, trust_card, metadata)
    _write_channels_sheet(wb, summary)
    _write_trust_card_sheet(wb, trust_card)
    _write_parameters_sheet(wb, summary)
    if refresh_diff is not None:
        _write_refresh_sheet(wb, refresh_diff)
    if scenarios:
        names = list(scenario_plan_names) if scenario_plan_names else []
        if len(names) < len(scenarios):
            names += [f"scenario_{i + 1}" for i in range(len(names), len(scenarios))]
        _write_scenarios_sheet(wb, list(scenarios), names[: len(scenarios)])
    return wb


def write_excel_workbook(
    summary: PosteriorSummary,
    trust_card: TrustCard,
    *,
    output: Path,
    metadata: WorkbookMetadata,
    refresh_diff: RefreshDiff | None = None,
    scenarios: list[ForecastResult] | None = None,
    scenario_plan_names: list[str] | None = None,
) -> Path:
    """Build the workbook and save it to ``output``. Returns ``output``."""
    wb = build_excel_workbook(
        summary,
        trust_card,
        metadata=metadata,
        refresh_diff=refresh_diff,
        scenarios=scenarios,
        scenario_plan_names=scenario_plan_names,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output
