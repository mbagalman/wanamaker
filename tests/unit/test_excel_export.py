"""Tests for the analyst-facing Excel workbook export.

These cover the structural promises of the workbook: which sheets are
present, that headline numbers match the inputs (channel contributions,
ROI, total media), that spend-invariant channels are flagged, that
optional sheets (Refresh diff, Scenarios) only appear when their inputs
are supplied, and that the file can be round-tripped via openpyxl.
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from wanamaker.engine.summary import (
    ChannelContributionSummary,
    ConvergenceSummary,
    ParameterSummary,
    PosteriorSummary,
    PredictiveSummary,
)
from wanamaker.forecast.posterior_predictive import (
    ExtrapolationFlag,
    ForecastResult,
)
from wanamaker.refresh.classify import MovementClass
from wanamaker.refresh.diff import ParameterMovement, RefreshDiff
from wanamaker.reports import (
    WorkbookMetadata,
    build_excel_workbook,
    write_excel_workbook,
)
from wanamaker.trust_card.card import TrustCard, TrustDimension, TrustStatus


def _channel(
    name: str,
    *,
    contribution: float = 5000.0,
    spend_invariant: bool = False,
    roi_mean: float = 2.0,
) -> ChannelContributionSummary:
    return ChannelContributionSummary(
        channel=name,
        mean_contribution=contribution,
        hdi_low=contribution * 0.9,
        hdi_high=contribution * 1.1,
        roi_mean=roi_mean,
        roi_hdi_low=roi_mean * 0.8,
        roi_hdi_high=roi_mean * 1.2,
        observed_spend_min=100.0,
        observed_spend_max=500.0,
        spend_invariant=spend_invariant,
    )


def _summary(
    *channels: ChannelContributionSummary,
    parameters: list[ParameterSummary] | None = None,
) -> PosteriorSummary:
    return PosteriorSummary(
        parameters=list(parameters) if parameters else [],
        channel_contributions=list(channels),
        convergence=ConvergenceSummary(
            max_r_hat=1.005,
            min_ess_bulk=500.0,
            n_divergences=0,
            n_chains=4,
            n_draws=1000,
        ),
        in_sample_predictive=PredictiveSummary(
            periods=["2024-01-01", "2024-06-01", "2024-12-31"],
            mean=[100.0, 100.0, 100.0],
            hdi_low=[90.0, 90.0, 90.0],
            hdi_high=[110.0, 110.0, 110.0],
        ),
    )


def _card(*dims: tuple[str, TrustStatus, str]) -> TrustCard:
    return TrustCard(
        dimensions=[
            TrustDimension(name=name, status=status, explanation=expl)
            for (name, status, expl) in dims
        ]
    )


def _metadata() -> WorkbookMetadata:
    return WorkbookMetadata(
        run_id="abc12345_20260101T000000Z",
        period_start="2024-01-01",
        period_end="2024-12-31",
        n_periods=52,
        generated_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
        package_version="0.1.0",
        runtime_mode="quick",
        engine_label="pymc 5.10.0",
        run_fingerprint="abc12345def67890",
    )


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------


def test_default_workbook_has_summary_channels_trust_card_sheets() -> None:
    summary = _summary(_channel("paid_search"))
    card = _card(("convergence", TrustStatus.PASS, "Clean."))

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    assert wb.sheetnames == ["Summary", "Channels", "Trust Card"]


def test_parameters_sheet_appears_when_parameters_present() -> None:
    summary = _summary(
        _channel("paid_search"),
        parameters=[
            ParameterSummary(
                name="channel.paid_search.ec50",
                mean=3000.0,
                sd=200.0,
                hdi_low=2600.0,
                hdi_high=3400.0,
                r_hat=1.001,
                ess_bulk=950.0,
            ),
        ],
    )
    card = _card(("convergence", TrustStatus.PASS, "Clean."))

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    assert "Parameters" in wb.sheetnames


def test_refresh_sheet_appears_only_when_diff_supplied() -> None:
    summary = _summary(_channel("paid_search"))
    card = _card(("convergence", TrustStatus.PASS, "Clean."))
    diff = RefreshDiff(
        previous_run_id="prev",
        current_run_id="curr",
        movements=[
            ParameterMovement(
                name="channel.paid_search.roi",
                previous_mean=2.0,
                current_mean=2.05,
                previous_ci=(1.8, 2.2),
                current_ci=(1.85, 2.25),
                movement_class=MovementClass.WITHIN_PRIOR_CI,
            ),
        ],
    )

    wb_no_diff = build_excel_workbook(summary, card, metadata=_metadata())
    wb_with_diff = build_excel_workbook(
        summary, card, metadata=_metadata(), refresh_diff=diff
    )

    assert "Refresh diff" not in wb_no_diff.sheetnames
    assert "Refresh diff" in wb_with_diff.sheetnames


def test_scenarios_sheet_appears_only_when_scenarios_supplied() -> None:
    summary = _summary(_channel("paid_search"))
    card = _card(("convergence", TrustStatus.PASS, "Clean."))
    forecast = ForecastResult(
        periods=["2026-01-05", "2026-01-12"],
        mean=[1000.0, 1100.0],
        hdi_low=[900.0, 980.0],
        hdi_high=[1100.0, 1220.0],
        extrapolation_flags=[],
        spend_invariant_channels=[],
    )

    wb_no = build_excel_workbook(summary, card, metadata=_metadata())
    wb_yes = build_excel_workbook(
        summary,
        card,
        metadata=_metadata(),
        scenarios=[forecast],
        scenario_plan_names=["q1_plan"],
    )

    assert "Scenarios" not in wb_no.sheetnames
    assert "Scenarios" in wb_yes.sheetnames


# ---------------------------------------------------------------------------
# Channel sheet content
# ---------------------------------------------------------------------------


def test_channels_sheet_one_row_per_channel_ranked_by_contribution() -> None:
    summary = _summary(
        _channel("small", contribution=1000.0),
        _channel("big", contribution=10_000.0),
        _channel("middle", contribution=5000.0),
    )
    card = _card()

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    ws = wb["Channels"]

    # Header + 3 rows.
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 4
    assert rows[0][0] == "Channel"
    assert rows[1][0] == "big"
    assert rows[2][0] == "middle"
    assert rows[3][0] == "small"


def test_channels_sheet_marks_spend_invariant() -> None:
    summary = _summary(
        _channel("active"),
        _channel("flat", spend_invariant=True),
    )
    card = _card()

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    ws = wb["Channels"]

    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    invariant_col = headers.index("Spend invariant")
    roi_col = headers.index("ROI (mean)")

    by_name = {row[0]: row for row in rows[1:]}
    assert by_name["active"][invariant_col] == "no"
    assert by_name["flat"][invariant_col] == "yes"
    # ROI cells for spend-invariant channels are blank, not zero.
    assert by_name["flat"][roi_col] is None
    assert by_name["active"][roi_col] is not None


def test_channels_sheet_share_of_media_is_a_proportion() -> None:
    summary = _summary(
        _channel("a", contribution=4000.0),
        _channel("b", contribution=1000.0),
    )
    card = _card()

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    ws = wb["Channels"]
    rows = list(ws.iter_rows(values_only=True))
    share_col = list(rows[0]).index("Share of media")
    by_name = {row[0]: row for row in rows[1:]}
    # 4000 / 5000 = 0.8 ; 1000 / 5000 = 0.2
    assert abs(by_name["a"][share_col] - 0.8) < 1e-9
    assert abs(by_name["b"][share_col] - 0.2) < 1e-9


# ---------------------------------------------------------------------------
# Summary sheet content
# ---------------------------------------------------------------------------


def test_summary_sheet_includes_total_media_and_verdict() -> None:
    summary = _summary(
        _channel("a", contribution=3000.0),
        _channel("b", contribution=2000.0),
    )
    card = _card(
        ("convergence", TrustStatus.PASS, "Clean."),
        ("saturation_identifiability", TrustStatus.WEAK, "Affiliate flat."),
    )

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    ws = wb["Summary"]

    label_to_value = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}

    assert label_to_value["Total media contribution"] == 5000.0
    assert label_to_value["Trust Card verdict"] == "weak"
    assert "a" in str(label_to_value["Top channel by contribution"])


# ---------------------------------------------------------------------------
# Trust Card sheet
# ---------------------------------------------------------------------------


def test_trust_card_sheet_one_row_per_dimension() -> None:
    summary = _summary(_channel("paid_search"))
    card = _card(
        ("convergence", TrustStatus.PASS, "Clean."),
        ("holdout_accuracy", TrustStatus.MODERATE, "Mixed."),
    )

    wb = build_excel_workbook(summary, card, metadata=_metadata())
    ws = wb["Trust Card"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 dimensions
    assert rows[1][0] == "convergence"
    assert rows[1][1] == "pass"
    assert rows[2][0] == "holdout_accuracy"
    assert rows[2][1] == "moderate"


# ---------------------------------------------------------------------------
# Round-trip
# ---------------------------------------------------------------------------


def test_workbook_round_trips_through_openpyxl(tmp_path: Path) -> None:
    summary = _summary(_channel("paid_search"))
    card = _card(("convergence", TrustStatus.PASS, "Clean."))

    output = tmp_path / "summary.xlsx"
    write_excel_workbook(summary, card, output=output, metadata=_metadata())

    assert output.exists()
    assert output.stat().st_size > 0

    wb = load_workbook(output, read_only=True)
    assert "Summary" in wb.sheetnames
    assert "Channels" in wb.sheetnames
    assert "Trust Card" in wb.sheetnames


# ---------------------------------------------------------------------------
# Scenarios sheet
# ---------------------------------------------------------------------------


def test_scenarios_sheet_delta_is_zero_for_first_plan_then_signed_for_others() -> None:
    summary = _summary(_channel("paid_search"))
    card = _card()

    def fr(mean: float) -> ForecastResult:
        return ForecastResult(
            periods=["2026-01-05", "2026-01-12"],
            mean=[mean, mean],
            hdi_low=[mean * 0.9, mean * 0.9],
            hdi_high=[mean * 1.1, mean * 1.1],
            extrapolation_flags=[
                ExtrapolationFlag(
                    period="2026-01-05",
                    channel="paid_search",
                    planned_spend=900.0,
                    observed_spend_min=100.0,
                    observed_spend_max=500.0,
                    direction="above",
                ),
            ],
            spend_invariant_channels=[],
        )

    wb = build_excel_workbook(
        summary,
        card,
        metadata=_metadata(),
        scenarios=[fr(1000), fr(1100), fr(900)],
        scenario_plan_names=["base", "alt_a", "alt_b"],
    )
    ws = wb["Scenarios"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    delta_col = headers.index("Δ vs first plan")
    extrap_col = headers.index("Extrapolation flags")

    # Row 1 = base, baseline; delta is 0.
    assert rows[1][delta_col] == 0.0
    # Row 2 = alt_a; total 2200 vs 2000 baseline, delta = +200.
    assert rows[2][delta_col] == 200.0
    # Row 3 = alt_b; total 1800 vs 2000 baseline, delta = -200.
    assert rows[3][delta_col] == -200.0
    # All scenarios have one extrapolation flag.
    assert rows[1][extrap_col] == 1
